from flask import Flask, render_template, request, redirect, url_for
import pandas as pd
from fuzzywuzzy import fuzz, process
from collections import defaultdict
from models import db, PayerGroup, Payer, PayerDetail
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload
import logging
import re
import os
from openpyxl import load_workbook
from rapidfuzz import fuzz as rapid_fuzz  # Faster alternative to fuzzywuzzy
import multiprocessing as mp
import pandas as pd
import logging
import os
from openpyxl import load_workbook
from multiprocessing import Pool, cpu_count
from functools import partial
import time
from sqlalchemy.sql import text
# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///payers.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

# Matching thresholds
THRESHOLDS = {
    'keyword_match': 80,
    'payer_number_match': 70,
    'semantic_match': 75,
    'dynamic_group_match': 80,
    'name_similarity': 85
}

# Predefined payer groups with aliases
KNOWN_GROUPS = {
    "Delta Dental": ["delta dental", "delta", "dd", "deltacare"],
    "Blue Cross Blue Shield": ["blue cross", "blue shield", "bcbs", "bluecross", "anthem", "wellpoint", "carefirst"],
    "Aetna": ["aetna", "aetna dental", "aetna life"],
    "Cigna": ["cigna", "cigna dental", "connecticut general"],
    "UnitedHealthcare": ["unitedhealthcare", "united healthcare", "uhc", "united health", "uhg", "optum"],
    "MetLife": ["metlife", "met life", "metropolitan"],
    "Medicare": ["medicare", "medicare advantage", "cms"],
    "Medicaid": ["medicaid", "medical assistance"]
}

IGNORED_WORDS = ["of", "in", "at", "for", "the", "and", "inc", "corp", "llc", "corporation", "company", "dental", "plan", "group", "insurance"]

# Initialize database
with app.app_context():
    db.create_all()

def process_sheet(args):
    file_path, sheet_name, required_columns = args
    try:
        logger.info(f"Processing sheet {sheet_name} in parallel")
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            engine='openpyxl',
            dtype={'Payer ID': str},
            usecols=lambda x: x in required_columns or x == 'Payer Name'
        )
        if df.empty:
            return None
        
        df = df.rename(columns={'Payer Name': 'Payer Identification Information'})
        if all(col in df.columns for col in required_columns):
            return df[required_columns].drop_duplicates()
        else:
            logger.warning(f"Sheet {sheet_name} missing required columns")
            return None
    except Exception as e:
        logger.error(f"Error processing sheet {sheet_name}: {e}")
        return None

# Optimized Excel loading with parallel processing
def load_excel_data(file_path=os.path.join('data', 'payers.xlsx')):
    start_time = time.time()
    try:
        # Identify non-legend sheets
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet_names = [name for name in wb.sheetnames if 'legend' not in name.lower()]
        wb.close()
        logger.info(f"Found {len(wb.sheetnames)} sheets, processing {len(sheet_names)} non-legend sheets: {sheet_names}")

        if not sheet_names:
            logger.warning("No non-legend sheets found")
            return pd.DataFrame()

        required_columns = ['Payer ID', 'Payer Identification Information']
        
        # Parallel processing of sheets
        with Pool(processes=min(cpu_count(), len(sheet_names))) as pool:
            process_func = partial(process_sheet, file_path=file_path, required_columns=required_columns)
            results = pool.map(process_func, [(file_path, sheet_name, required_columns) for sheet_name in sheet_names])

        # Combine results
        combined_df = pd.concat([df for df in results if df is not None], ignore_index=True)
        
        if not combined_df.empty:
            # Final deduplication
            combined_df = combined_df.drop_duplicates(subset=required_columns)
            logger.info(f"Loaded {len(combined_df)} rows in {time.time() - start_time:.2f} seconds")
        else:
            logger.info("No data loaded after processing")

        return combined_df

    except Exception as e:
        logger.error(f"Error loading Excel data: {e}")
        return pd.DataFrame()

# Extract key terms (vectorized where possible)
def extract_key_terms(payer_name):
    if not isinstance(payer_name, str):
        return [], ""
    name = payer_name.lower().strip()
    name = re.sub(r'\b(inc|llc|corp)\b|[.,()]', '', name)
    words = name.split()
    key_terms = [word for word in words if word not in IGNORED_WORDS and len(word) > 2]
    return key_terms, " ".join(key_terms)

# Batch infer payer groups
def infer_payer_group_batch(rows, existing_groups):
    results = []
    for payer_name, payer_number in rows:
        if not isinstance(payer_name, str) or pd.isna(payer_name):
            results.append("Unknown")
            continue
        payer_name = payer_name.lower().strip()
        key_terms, simplified_name = extract_key_terms(payer_name)
        
        # Step 1: Match with known groups
        for group_name, keywords in KNOWN_GROUPS.items():
            if any(rapid_fuzz.partial_ratio(keyword, payer_name) > THRESHOLDS['keyword_match'] for keyword in keywords):
                results.append(group_name)
                break
        else:
            # Step 2: Match by payer number
            if payer_number:
                for group_name, payers in existing_groups.items():
                    for payer in payers:
                        if payer.get("payer_number") == payer_number and rapid_fuzz.ratio(payer["name"], payer_name) > THRESHOLDS['payer_number_match']:
                            results.append(group_name)
                            break
                    else:
                        continue
                    break
                else:
                    # Step 3: Semantic matching
                    for group_name in existing_groups:
                        if rapid_fuzz.partial_ratio(group_name.lower(), simplified_name) > THRESHOLDS['semantic_match']:
                            results.append(group_name)
                            break
                    else:
                        results.append(simplified_name.title() if simplified_name else "Unknown")
            else:
                results.append(simplified_name.title() if simplified_name else "Unknown")
    return results

# Optimized mapping algorithm
def map_payer_details(df):
    # Load existing data efficiently
    existing_groups = defaultdict(list)
    payers_by_group = db.session.query(PayerGroup).options(joinedload(PayerGroup.payers).joinedload(Payer.payer_details)).all()
    for group in payers_by_group:
        for payer in group.payers:
            for detail in payer.payer_details:
                existing_groups[group.name].append({
                    "name": payer.name,
                    "payer_number": detail.payer_number
                })

    # Batch process group inference
    rows = list(zip(df['Payer Identification Information'], df['Payer ID'].fillna('')))
    with mp.Pool(processes=mp.cpu_count()) as pool:
        group_names = pool.apply(infer_payer_group_batch, (rows, existing_groups))
    
    # Pre-fetch existing groups and payers
    group_map = {g.name: g for g in PayerGroup.query.all()}
    payer_map = {(p.payer_group_id, p.name): p for p in Payer.query.all()}
    
    # Batch insertions
    new_groups = []
    new_payers = []
    new_details = []
    
    for (payer_name, payer_id), group_name in zip(rows, group_names):
        payer_id = payer_id if pd.notna(payer_id) else None
        
        # Handle group
        if group_name not in group_map:
            new_groups.append(PayerGroup(name=group_name))
            group_map[group_name] = new_groups[-1]
        
        group = group_map[group_name]
        
        # Handle payer
        payer_key = (group.id if group.id else None, payer_name)
        if payer_key not in payer_map:
            new_payers.append(Payer(name=payer_name, payer_group_id=group.id))
            payer_map[payer_key] = new_payers[-1]
        
        payer = payer_map[payer_key]
        new_details.append(PayerDetail(payer_id=payer.id if payer.id else None, payer_name_raw=payer_name, payer_number=payer_id))
    
    # Bulk insert
    if new_groups:
        db.session.bulk_save_objects(new_groups)
        db.session.commit()
        for g in new_groups:
            group_map[g.name] = PayerGroup.query.filter_by(name=g.name).first()
    
    if new_payers:
        for p in new_payers:
            p.payer_group_id = group_map[p.payer_group.name].id
        db.session.bulk_save_objects(new_payers)
        db.session.commit()
    
    if new_details:
        for d in new_details:
            d.payer_id = payer_map[(group_map[group_names[rows.index((d.payer_name_raw, d.payer_number or ''))]].id, d.payer_name_raw)].id
        db.session.bulk_save_objects(new_details)
    
    db.session.commit()
    logger.info("Mapping completed")
def get_all_groups():
    """
    Retrieve all payer groups from the database.
    
    Returns:
        list: A list of PayerGroup objects, each representing a unique payer group.
    """
    try:
        groups = PayerGroup.query.all()
        logger.info(f"Retrieved {len(groups)} payer groups from the database")
        return groups
    except Exception as e:
        logger.error(f"Error retrieving payer groups: {e}")
        return []
# Routes (unchanged for brevity)
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/load_data', methods=['POST'])
def load_data():
    df = load_excel_data()
    if not df.empty:
        map_payer_details(df)
    return redirect(url_for('map_payers'))

@app.route('/map_payers')
def map_payers():
    payer_details = PayerDetail.query.all()
    return render_template('map_payers.html', payer_details=payer_details)

@app.route('/manage_groups', methods=['GET', 'POST'])
def manage_groups():
    if request.method == 'POST':
        source_group_id = request.form.get('source_group_id')
        target_group_id = request.form.get('target_group_id')
        source_group = PayerGroup.query.get(source_group_id)
        target_group = PayerGroup.query.get(target_group_id)
        if source_group and target_group:
            for payer in source_group.payers:
                payer.payer_group_id = target_group.id
            db.session.delete(source_group)
            db.session.commit()
        return redirect(url_for('manage_groups'))
    groups = PayerGroup.query.all()
    return render_template('manage_groups.html', groups=groups)
@app.route('/merge_groups', methods=['POST'])
def merge_groups():
    source_id = request.form.get('source_group')
    target_id = request.form.get('target_group')
    
    if not source_id or not target_id or source_id == target_id:
        return render_template('manage_groups.html', 
                              error="Invalid source or target group selection", 
                              groups=get_all_groups())
    
    try:
        # Start a transaction
        with db.session.begin():
            # Make sure target_id is valid and exists
            target_group = db.session.query(PayerGroup).get(target_id)
            if not target_group:
                raise ValueError("Target group does not exist")
                
            # Update all payers from source group to target group
            db.session.execute(
                text("UPDATE payer SET payer_group_id = :target_id WHERE payer_group_id = :source_id"),
                {"source_id": source_id, "target_id": target_id}
            )
            
            # Optionally delete the source group after merging
            db.session.execute(
                text("DELETE FROM payer_group WHERE id = :source_id"),
                {"source_id": source_id}
            )
        
        return redirect(url_for('manage_groups'))
    except Exception as e:
        db.session.rollback()
        return render_template('manage_groups.html', 
                              error=f"Error during merge: {str(e)}", 
                              groups=get_all_groups())
@app.route('/set_pretty_name/<int:payer_id>', methods=['GET', 'POST'])
def set_pretty_name(payer_id):
    payer = Payer.query.get_or_404(payer_id)
    if request.method == 'POST':
        pretty_name = request.form['pretty_name']
        payer.pretty_name = pretty_name
        db.session.commit()
        return redirect(url_for('map_payers'))
    return render_template('set_pretty_name.html', payer=payer)

if __name__ == '__main__':
    app.run(debug=True)